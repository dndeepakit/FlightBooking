import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from datetime import datetime

st.set_page_config(page_title="Flight Options Tool", layout="wide")

st.title("✈️ Flight Options Generator (Refined Version)")
st.write("Generate professional flight comparison tables for employee travel planning.")

# ---- INPUT SECTION ----
st.header("🔹 Trip Details")

col1, col2, col3, col4 = st.columns(4)
with col1:
    dep_city = st.text_input("Departure City", "")
with col2:
    arr_city = st.text_input("Arrival City", "")
with col3:
    dep_date = st.date_input("Departure Date")
with col4:
    ret_date = st.date_input("Return Date")

st.divider()

# ---- ADD FLIGHT OPTIONS ----
st.header("🛫 Flight Options")

num_options = st.number_input("Number of Options", min_value=1, max_value=20, value=3)
data = []

for i in range(num_options):
    st.subheader(f"Option {i+1}")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        outbound_airline = st.text_input(f"Outbound Airline (Option {i+1})", "")
    with c2:
        outbound_dep_time = st.text_input(f"Outbound Departure Time (Option {i+1})", "")
    with c3:
        outbound_arr_time = st.text_input(f"Outbound Arrival Time (Option {i+1})", "")
    with c4:
        outbound_price = st.text_input(f"Round Trip Cost (Option {i+1})", "")

    c5, c6, c7, c8 = st.columns(4)
    with c5:
        return_airline = st.text_input(f"Return Airline (Option {i+1})", "")
    with c6:
        return_dep_time = st.text_input(f"Return Departure Time (Option {i+1})", "")
    with c7:
        return_arr_time = st.text_input(f"Return Arrival Time (Option {i+1})", "")
    with c8:
        _ = st.text_input(f"(Optional) Notes (Option {i+1})", "")

    data.append({
        "Option": f"Option {i+1}",
        "Outbound Airline": outbound_airline,
        "Return Airline": return_airline,
        "Outbound Departure Time": outbound_dep_time,
        "Outbound Arrival Time": outbound_arr_time,
        "Return Departure Time": return_dep_time,
        "Return Arrival Time": return_arr_time,
        "Round Trip Cost": outbound_price
    })

st.divider()

# ---- GENERATE OUTPUT ----
if st.button("Generate Flight Table"):
    if not dep_city or not arr_city:
        st.error("Please fill in both Departure and Arrival Cities.")
    else:
        df = pd.DataFrame(data)

        st.success("✅ Flight Options Generated Successfully!")
        st.subheader(f"Flight Options For {dep_date.strftime('%d %b %Y')} - {ret_date.strftime('%d %b %Y')}")

        st.dataframe(df, use_container_width=True)

        # ---- Excel generation ----
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Flight Options"

        # Styles
        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # Header
        title = f"Flight Options For {dep_date.strftime('%d %b %Y')} - {ret_date.strftime('%d %b %Y')}"
        ws.merge_cells('A1:H1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=13)
        ws['A1'].alignment = center

        # Table headers
        headers = ["Option", "Outbound Airline", "Return Airline",
                   "Outbound Departure Time", "Outbound Arrival Time",
                   "Return Departure Time", "Return Arrival Time", "Round Trip Cost"]

        ws.append(headers)

        for cell in ws[2]:
            cell.font = bold
            cell.alignment = center
            cell.border = border

        # Data rows
        for _, row in df.iterrows():
            ws.append(row.tolist())

        for row in ws.iter_rows(min_row=3, max_col=8):
            for cell in row:
                cell.alignment = center
                cell.border = border

        # Auto column width
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(buffer)
        buffer.seek(0)

        st.download_button(
            label="📥 Download Excel File",
            data=buffer,
            file_name=f"Flight_Options_{dep_city}_to_{arr_city}_{dep_date.strftime('%d%b')}-{ret_date.strftime('%d%b%Y')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # ---- Copyable Text Output ----
        st.subheader("📝 Copyable Text Summary")
        text_output = f"Flight Options For {dep_date.strftime('%d %b %Y')} - {ret_date.strftime('%d %b %Y')}\n"
        text_output += "-" * 80 + "\n\n"

        for i, row in df.iterrows():
            text_output += (
                f"{row['Option']}\n"
                f"OUTBOUND ({dep_date.strftime('%d/%m/%Y')}) | {row['Outbound Airline']} | DEP: {dep_city} {row['Outbound Departure Time']} | ARR: {arr_city} {row['Outbound Arrival Time']}\n"
                f"RETURN ({ret_date.strftime('%d/%m/%Y')}) | {row['Return Airline']} | DEP: {arr_city} {row['Return Departure Time']} | ARR: {dep_city} {row['Return Arrival Time']}\n"
                f"Round Trip COST: ₹{row['Round Trip Cost']}\n\n"
            )

        st.text_area("Copy this formatted text:", text_output, height=300)

        st.subheader("🌍 Quick Search Links")
        st.markdown(f"""
        - [MakeMyTrip](https://www.makemytrip.com/flights)
        - [Skyscanner](https://www.skyscanner.co.in/)
        - [EaseMyTrip](https://www.easemytrip.com/)
        - [Goibibo](https://www.goibibo.com/)
        - [ClearTrip](https://www.cleartrip.com/flights)
        - [Yatra](https://www.yatra.com/flights)
        """)

st.markdown("---")
st.caption("© 2025 Flight Options Tool | Cloud-based version for Admin Team")
